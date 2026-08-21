"""Exportação do dossiê consolidado em Excel (4 abas)."""

from __future__ import annotations

import io
from collections.abc import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..core.cnpj import formatar as formatar_cnpj
from ..core.formatters import url_google_maps
from ..core.models import Empresa

_FILL_TITULO = PatternFill("solid", start_color="1F497D", end_color="1F497D")
_FILL_HEADER = PatternFill("solid", start_color="2F5597", end_color="2F5597")
_FONTE_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_FONTE_CORPO = Font(name="Calibri", size=11)
_BORDA = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
_LARGURA_MAX = 60


def _cabecalho(ws: Worksheet, headers: Sequence[str], linha: int = 1) -> None:
    ws.row_dimensions[linha].height = 24
    for col, titulo in enumerate(headers, 1):
        cel = ws.cell(row=linha, column=col, value=titulo)
        cel.font = _FONTE_HEADER
        cel.fill = _FILL_HEADER
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)
    ws.auto_filter.ref = (
        f"A{linha}:{get_column_letter(len(headers))}{linha}"
    )


def _escrever(ws: Worksheet, linha: int, valores: Sequence) -> None:
    for col, valor in enumerate(valores, 1):
        cel = ws.cell(row=linha, column=col, value=valor)
        cel.font = _FONTE_CORPO
        cel.border = _BORDA
        cel.alignment = Alignment(vertical="top", wrap_text=isinstance(valor, str)
                                  and len(valor) > 40)


def _ajustar_larguras(ws: Worksheet, linha_header: int = 1) -> None:
    """[MELHORIA] O original importava ``get_column_letter`` e nunca usava —
    todas as colunas saíam com a largura padrão, cortando o texto."""
    for col in range(1, ws.max_column + 1):
        maior = 0
        for linha in range(linha_header, min(ws.max_row, 400) + 1):
            valor = ws.cell(row=linha, column=col).value
            if valor is not None:
                maior = max(maior, min(len(str(valor)), _LARGURA_MAX))
        ws.column_dimensions[get_column_letter(col)].width = max(12, maior + 2)


def gerar_dossie_excel(empresas: Sequence[Empresa]) -> bytes:
    """Devolve ``bytes`` (não ``BytesIO``) — evita o bug de ponteiro no fim do
    buffer quando o mesmo objeto é reaproveitado em mais de um download."""
    wb = Workbook()

    # ---------------- ABA 1: Resumo cadastral ---------------------------- #
    ws1 = wb.active
    ws1.title = "Resumo Cadastral"
    ws1.merge_cells("A1:O1")
    titulo = ws1["A1"]
    titulo.value = "DOSSIÊ DE ONBOARDING CONTÁBIL"
    titulo.font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    titulo.fill = _FILL_TITULO
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    _cabecalho(ws1, [
        "CNPJ", "Razão Social", "Nome Fantasia", "Situação", "Regime", "E-mail",
        "Telefone", "CNAE Principal", "Anexo Simples", "Capital Social",
        "Inscr. Municipal", "Inscr. Estadual", "Endereço", "Cód. IBGE", "Google Maps",
    ], linha=3)

    for i, emp in enumerate(empresas, 4):
        _escrever(ws1, i, [
            formatar_cnpj(emp.cnpj), emp.razao_social, emp.nome_fantasia,
            emp.situacao.situacao_receita, emp.regime, emp.email_str, emp.telefone_str,
            emp.cnae_principal_str,
            emp.atividade_principal.diagnostico.anexo if emp.atividade_principal else "",
            emp.capital_social, emp.inscricao_municipal,
            ", ".join(emp.inscricoes_estaduais) or "Isento / não informado",
            emp.endereco.linha_completa, emp.endereco.cod_ibge,
            url_google_maps(emp.endereco.linha_completa),
        ])
        ws1.cell(row=i, column=10).number_format = 'R$ #,##0.00'
    _ajustar_larguras(ws1, linha_header=3)

    # ---------------- ABA 2: Análise tributária -------------------------- #
    ws2 = wb.create_sheet("Análise Tributária")
    _cabecalho(ws2, ["CNPJ", "Razão Social", "Tipo", "CNAE", "Descrição",
                     "Anexo", "Alíquota", "Diagnóstico / Oportunidade"])
    linha = 2
    for emp in empresas:
        atividades = []
        if emp.atividade_principal:
            atividades.append(("PRINCIPAL", emp.atividade_principal))
        atividades += [("SECUNDÁRIO", a) for a in emp.atividades_secundarias]
        for tipo, ativ in atividades:
            _escrever(ws2, linha, [
                formatar_cnpj(emp.cnpj), emp.razao_social, tipo, ativ.codigo,
                ativ.descricao, ativ.diagnostico.anexo,
                ativ.diagnostico.aliquota_inicial, ativ.diagnostico.dica_engenharia,
            ])
            linha += 1
    _ajustar_larguras(ws2)

    # ---------------- ABA 3: Compliance ---------------------------------- #
    ws3 = wb.create_sheet("Compliance")
    _cabecalho(ws3, ["CNPJ", "Razão Social", "Situação na Receita",
                     "Data da Situação", "Pendente de verificação formal"])
    for i, emp in enumerate(empresas, 2):
        _escrever(ws3, i, [
            formatar_cnpj(emp.cnpj), emp.razao_social, emp.situacao.rotulo_receita,
            emp.situacao.data_situacao or "n/d",
            " | ".join(emp.situacao.pendentes_de_verificacao),
        ])
    _ajustar_larguras(ws3)

    # ---------------- ABA 4: QSA ----------------------------------------- #
    ws4 = wb.create_sheet("Quadro Societário")
    _cabecalho(ws4, ["CNPJ", "Razão Social", "Sócio / Administrador",
                     "Qualificação", "Faixa Etária", "Alerta"])
    linha = 2
    for emp in empresas:
        if emp.socios:
            for socio in emp.socios:
                _escrever(ws4, linha, [
                    formatar_cnpj(emp.cnpj), emp.razao_social, socio.nome,
                    socio.qualificacao, socio.faixa_etaria,
                    "⚠️ Verificar participação ≥10% em outra empresa do Simples"
                    if emp.tem_risco_societario else "Sócio único",
                ])
                linha += 1
        else:
            _escrever(ws4, linha, [
                formatar_cnpj(emp.cnpj), emp.razao_social,
                "Empresário Individual / MEI", "N/A", "N/A", "Sem sócios",
            ])
            linha += 1
    _ajustar_larguras(ws4)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
